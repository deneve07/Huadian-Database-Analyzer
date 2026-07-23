"""
華典資料庫分析系統 (Streamlit 版)
----------------------------------
1. 自訂樞紐分析 - 用拖拉的方式，從各資料集「所有欄位」中自由挑選、排序要放進報表的欄位，
   每個欄位可直接篩選、決定是否合計，並可加入占比(%)、成長率(%) 計算欄位。
2. 處方釋出率分析 - 固定格式，維持原本的計算邏輯與版面，不提供自訂欄位。

需要的套件（requirements.txt）：
    streamlit
    pandas
    numpy
    openpyxl
    pyarrow
    streamlit-sortables
"""

import re
import io
import os
import json
import base64
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
    TAIPEI_TZ = ZoneInfo("Asia/Taipei")
except Exception:
    TAIPEI_TZ = None
import pandas as pd
import numpy as np
import streamlit as st
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

try:
    from streamlit_sortables import sort_items
    HAS_SORTABLES = True
except ImportError:
    HAS_SORTABLES = False


# ============================================================
# 基礎設定與資料載入
# ============================================================

st.set_page_config(page_title="華典資料庫分析系統", page_icon="📊", layout="wide")


@st.cache_data(show_spinner=False)
def load_data(path: str) -> pd.DataFrame:
    parquet_path = path.rsplit(".", 1)[0] + ".parquet"
    try:
        if os.path.exists(parquet_path):
            df = pd.read_parquet(parquet_path)
        else:
            # utf-8-sig 可自動去除 Excel 匯出 CSV 常見的 BOM 字元，
            # 否則第一欄欄名容易變成「\ufeff成分簡稱」導致 KeyError
            # keep_default_na=False, na_filter=False：避免 pandas 讀取階段就把 "#N/A"、"N/A"、
            # "NA"、"NULL"、"None" 等字串自動判定為空值並轉成 NaN (即使 dtype=str 也一樣會轉)，
            # 導致這些原始文字在我們自訂的缺值/異常文字判斷邏輯執行之前就已經遺失、
            # 沒辦法跟真正的空白儲存格或 "-" 佔位符做出區分。改為手動控制何謂缺值。
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", keep_default_na=False, na_filter=False)
    except Exception as e:
        # 靜默吞掉讀取錯誤會讓問題延遲到後面 (例如樞紐分析階段) 才爆成難以追查的
        # KeyError；改為明確顯示讀取失敗原因，方便直接定位是檔案格式的問題
        st.error(f"⚠️ 讀取資料檔案「{path}」時發生錯誤，已略過此檔案：{e}")
        return pd.DataFrame()

    df.columns = [str(c).strip() for c in df.columns]

    if "劑型小分類" in df.columns:
        df = df.rename(columns={"劑型小分類": "劑型"})
    rename_map = {}
    for y in ["2022", "2023", "2024"]:
        src = f"{y}年數量(顆)"
        if src in df.columns:
            rename_map[src] = f"{y}年申報量(顆)"
    if rename_map:
        df = df.rename(columns=rename_map)

    # 偵測重複欄位名稱：若來源檔案 (或改名後) 出現同名欄位，df[col] 會回傳
    # DataFrame 而非單一 Series，後續加總/格式化會出現非預期的型別，
    # 曾實際導致報表格式化時崩潰。及早警示、合併重複欄位 (取加總) 避免此問題。
    if df.columns.duplicated().any():
        dup_names = sorted(set(df.columns[df.columns.duplicated()]))
        st.warning(
            f"⚠️ 資料檔案裡有重複的欄位名稱：{', '.join(dup_names)}，"
            f"已自動合併 (數字欄位加總、文字欄位取第一筆)，請確認來源檔案是否有誤。"
        )
        merged = {}
        for name in dict.fromkeys(df.columns):
            same = df.loc[:, df.columns == name]
            if same.shape[1] == 1:
                merged[name] = same.iloc[:, 0]
            elif ("申報量" in name) or ("金額" in name):
                # 加總前務必先清除千分位逗號/空白/缺值佔位符，否則會重蹈
                # 「帶逗號的大數字被 pd.to_numeric 誤判成 NaN 而變成 0」的覆轍
                cleaned_cols = same.apply(
                    lambda s: s.astype(str).str.strip().str.replace(",", "", regex=False).str.replace(" ", "", regex=False)
                )
                blank_vals = {"", "nan", "none", "-", "--", "－", "null", "無"}
                cleaned_cols = cleaned_cols.apply(lambda s: s.mask(s.str.lower().isin(blank_vals), None))
                merged[name] = cleaned_cols.apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)
            else:
                merged[name] = same.iloc[:, 0]
        df = pd.DataFrame(merged)

    for col in df.columns:
        if ("申報量" in col) or ("金額" in col):
            # 嚴重錯誤修正：來源 CSV 的數字欄位可能帶有千分位逗號 (例如 "34,490,060")
            # 或前後空白，pd.to_numeric() 遇到逗號會直接判定為無法轉換而回傳 NaN，
            # 原本的 .fillna(0) 會把這些「其實是合法數字、只是格式帶逗號」的儲存格
            # 靜默吃成 0 —— 數字越大 (>=1000，含千分位逗號) 越容易中招，
            # 曾實際造成金額欄位大量歸零、特定醫院整列申報量消失的嚴重資料錯誤。
            cleaned = (
                df[col].astype(str)
                .str.strip()
                .str.replace(",", "", regex=False)
                .str.replace(" ", "", regex=False)
            )
            # 「無資料」佔位符號 (空白、"-"、全形"－"、"--"、"nan"、"無" 等)：
            # 加總計算一律當 0，但在報表明細列會顯示為「-」，跟「真正的0」做出區別，
            # 讓使用者能一眼看出這一格究竟是「本來就是0」還是「根本沒有申報資料」。
            blank_placeholder_values = {"", "nan", "none", "-", "--", "－", "null", "無"}
            cleaned_lower = cleaned.str.lower()
            is_blank_placeholder = cleaned_lower.isin(blank_placeholder_values)
            cleaned_for_numeric = cleaned.mask(is_blank_placeholder, None)
            numeric = pd.to_numeric(cleaned_for_numeric, errors="coerce")

            # 找出「不是空白/-類佔位符，但清理後仍無法轉成數字」的內容 (例如 "#N/A"、"Unknown"、
            # "NoN" 這類具資訊性的異常標記)。這種情況不當作單純缺值，加總計算依然視為 0，
            # 但在報表明細列會保留「原樣文字」顯示，讓使用者知道這裡不是掛零、而是資料本身
            # 標記了特殊狀態，同時仍跳出警告提醒人工確認來源資料。
            original_stripped = df[col].astype(str).str.strip()
            is_bad_text = original_stripped.notna() & (~is_blank_placeholder) & numeric.isna()

            # 顯示標記欄位：""=正常數字 (直接顯示格式化後的數字)；"-"=無資料佔位符；
            # 其餘字串=保留原樣顯示的異常標記文字 (例如 "#N/A"、"Unknown")
            marker = pd.Series("", index=df.index, dtype=object)
            marker[is_blank_placeholder] = "-"
            marker[is_bad_text] = original_stripped[is_bad_text]
            df[col + "__mk"] = marker

            if is_bad_text.any():
                bad_samples = df.loc[is_bad_text, col].astype(str).unique()[:5]
                st.warning(
                    f"⚠️ 欄位「{col}」有 {is_bad_text.sum()} 筆資料無法解析為數字，"
                    f"加總計算已視為 0，報表中會保留原樣文字顯示 (範例值：{', '.join(bad_samples)})"
                )

            df[col] = numeric.fillna(0)

    return df


def parse_dosage(d):
    m = re.search(r"[\d.]+", str(d))
    return float(m.group()) if m else 0.0


def order_display_cols(value_cols, pct_cols, growth_cols):
    """Req 4：占比要接在申報量後面 (金額之前)，而不是放在所有數值欄位最後"""
    qty_cols = [c for c in value_cols if "申報量" in c]
    other_cols = [c for c in value_cols if c not in qty_cols]
    return qty_cols + pct_cols + other_cols + growth_cols


def pretty_header(col: str):
    """把 '2022年申報量(顆)' 拆成 ('2022年','申報量')；'2022-2023年成長率(%)' 拆成 ('2022-2023年','成長率(%)')"""
    m = re.match(r"^(\d{4}(-\d{4})?年)(.+)$", col)
    if m:
        year = m.group(1)
        label = m.group(3).replace("(顆)", "")
        return year, label
    return None, col


# ============================================================
# 通用樞紐分析引擎：任意欄位順序 + 任意層級巢狀合計 + 占比/成長率
# ============================================================

def build_nested_rows(df: pd.DataFrame, row_fields: list, subtotal_fields: list, value_cols: list,
                       pct_years: list = None, add_growth: bool = False):
    # 顯示標記欄位 (由 load_data 產生)：""=正常數字、"-"=無資料佔位符、其他文字=異常標記 (如 #N/A、Unknown)。
    # 同一個報表欄位組合 (row_fields) 可能對應到來源資料裡的多筆原始紀錄，合併規則如下：
    # 只要其中有任一筆是「正常數字」，代表加總後這一格具有實際數字意義，優先顯示數字；
    # 全部都是「無資料」佔位符才顯示 "-"；若混雜了不同的異常文字則無法用單一文字代表，
    # 一律退回顯示數字 (此時應為 0)。這一步僅影響「顯示」，加總計算 (下面的 .sum()) 完全不受影響。
    #
    # 注意：標記結果改用獨立的查詢表 (dict) 儲存，刻意不用 merge() 併回主要的 df，
    # 避免合併後欄位結構跑掉、導致後面 sort_values() 等依賴欄位名稱的邏輯找不到欄位
    # (曾實際發生 KeyError)。這樣主流程的 df 從頭到尾維持跟修改前完全一樣的欄位結構。
    col_to_mk = {c: c + "__mk" for c in value_cols if c + "__mk" in df.columns}
    overrides_lookup = {}

    def combine_markers(vals):
        vals = list(dict.fromkeys(vals))  # 去重但保留順序，避免 set 的隨機順序
        if len(vals) == 1:
            return vals[0]
        if "" in vals:
            return ""
        if all(v == "-" for v in vals):
            return "-"
        return ""

    if col_to_mk:
        marker_cols = list(col_to_mk.values())
        grouped_marks = df.groupby(row_fields, as_index=False)[marker_cols].agg(
            lambda s: combine_markers(s.tolist())
        )
        mk_to_col = {mk: c for c, mk in col_to_mk.items()}
        for _, r in grouped_marks.iterrows():
            key = tuple(r[f] for f in row_fields)
            overrides = {}
            for mk_col in marker_cols:
                v = r[mk_col]
                if v:
                    overrides[mk_to_col[mk_col]] = v
            if overrides:
                overrides_lookup[key] = overrides

    # 先依「報表欄位」的組合彙總數值，確保同一個欄位組合 (例如同一家廠商) 只會出現一列，
    # 而不是把資料集裡每一筆原始紀錄 (可能還有藥品名稱、藥證字號等未顯示的欄位差異) 都個別列出來
    df = df.groupby(row_fields, as_index=False)[value_cols].sum()

    qty_cols = [c for c in value_cols if "申報量" in c]
    qty_years = sorted(set(c[:4] for c in qty_cols))
    pct_years = pct_years or []

    pct_cols = [f"{y}年占比(%)" for y in qty_years if y in pct_years]
    growth_cols = [f"{qty_years[i]}-{qty_years[i+1]}年成長率(%)" for i in range(len(qty_years) - 1)] if add_growth else []

    def compute_extra_for_row(sums, top_totals):
        extra = {}
        for c in pct_cols:
            y = c[:4]
            base = f"{y}年申報量(顆)"
            extra[c] = sums.get(base, 0) / top_totals.get(base, 0) if top_totals.get(base, 0) else 0
        for c in growth_cols:
            y1 = c[:4]
            y2 = c[5:9]
            b1, b2 = f"{y1}年申報量(顆)", f"{y2}年申報量(顆)"
            extra[c] = (sums.get(b2, 0) - sums.get(b1, 0)) / sums.get(b1, 0) if sums.get(b1, 0) else 0
        return extra

    def compute_extra_for_group(totals):
        extra = {c: 1.0 for c in pct_cols}  # 小計/總計列本身佔比恆為 100%
        for c in growth_cols:
            y1 = c[:4]
            y2 = c[5:9]
            b1, b2 = f"{y1}年申報量(顆)", f"{y2}年申報量(顆)"
            extra[c] = (totals.get(b2, 0) - totals.get(b1, 0)) / totals.get(b1, 0) if totals.get(b1, 0) else 0
        return extra

    year_cols = sorted(qty_cols, reverse=True)

    sort_cols = []
    sort_asc = []
    temp_sort_cols = []
    for col in subtotal_fields:
        if col == "含量":
            # 含量欄位需依實際劑量數值排序 (例如 12mg/ml < 24mg/ml < 100mg/ml)，
            # 而非把它當一般文字做字母排序 (那樣 "100mg/ml" 會排到 "12mg/ml" 前面)
            tmp_col = "__sortkey_含量"
            df[tmp_col] = df[col].map(parse_dosage)
            sort_cols.append(tmp_col)
            temp_sort_cols.append(tmp_col)
        else:
            sort_cols.append(col)
        sort_asc.append(True)
    sort_cols += year_cols
    sort_asc += [False] * len(year_cols)

    df_sorted = df.sort_values(by=sort_cols, ascending=sort_asc) if sort_cols else df
    if temp_sort_cols:
        df_sorted = df_sorted.drop(columns=temp_sort_cols)

    if subtotal_fields:
        min_level_idx = min(row_fields.index(c) for c in subtotal_fields)
    else:
        min_level_idx = len(row_fields)

    group_cols_to_blank = []
    for c in row_fields:
        if c in subtotal_fields:
            continue
        if df[c].nunique() <= 1 or row_fields.index(c) < min_level_idx:
            group_cols_to_blank.append(c)

    rows = []
    last_vals = {}
    first_flags = {c: True for c in subtotal_fields}

    def emit_row(row, top_totals):
        rec_vals = {}
        for f in row_fields:
            val = row[f]
            show = val
            if f in group_cols_to_blank:
                if val != last_vals.get(f):
                    last_vals[f] = val
                else:
                    show = ""
            elif f in first_flags:
                if not first_flags[f]:
                    show = ""
            rec_vals[f] = show
        sums = {c: row[c] for c in value_cols}
        sums.update(compute_extra_for_row(sums, top_totals))
        # 只有明細列 (data) 才會保留「-」或原樣異常文字的顯示標記；
        # 小計/總計列一律是多筆明細加總而成，即使全部都是 0 也視為正常數字顯示，不套用標記
        key = tuple(row[f] for f in row_fields)
        overrides = overrides_lookup.get(key, {})
        rows.append({"type": "data", "values": rec_vals, "sums": sums, "overrides": overrides})
        for c in first_flags:
            first_flags[c] = False

    def emit_subtotal(field, label, totals):
        rec_vals = {f: "" for f in row_fields}
        rec_vals[field] = f"{label} 合計"
        full = dict(totals)
        full.update(compute_extra_for_group(totals))
        rows.append({"type": "subtotal", "values": rec_vals, "sums": full})

    def recurse(sub_df, level_idx, top_totals):
        if level_idx >= len(subtotal_fields):
            for _, row in sub_df.iterrows():
                emit_row(row, top_totals)
            return
        col = subtotal_fields[level_idx]
        for _, grp in sub_df.groupby(col, sort=False):
            totals = {c: grp[c].sum() for c in value_cols}
            next_top = totals if level_idx == len(subtotal_fields) - 1 else top_totals
            label = grp[col].iloc[0]
            first_flags[col] = True
            recurse(grp, level_idx + 1, next_top)
            emit_subtotal(col, label, totals)

    grand_totals = {c: df[c].sum() for c in value_cols}
    if subtotal_fields:
        recurse(df_sorted, 0, grand_totals)
    else:
        for _, row in df_sorted.iterrows():
            emit_row(row, grand_totals)

    total_vals = {f: "" for f in row_fields}
    if row_fields:
        total_vals[row_fields[0]] = "總計"
    full_grand = dict(grand_totals)
    full_grand.update(compute_extra_for_group(grand_totals))
    rows.append({"type": "total", "values": total_vals, "sums": full_grand})

    return rows, pct_cols, growth_cols


def get_report_timestamp() -> str:
    """回傳報表產出當下的時間戳記字串 (24小時制、不含AM/PM)，格式例如 2026/7/23 10:28。
    刻意在產出當下就寫死成固定文字，而不是用 Excel 的 &D/&T 動態欄位代碼，
    因為動態代碼的實際顯示格式會依開啟檔案當下的 Excel 地區/語言設定而不同 (可能跑出
    12小時制加 AM/PM)，寫死文字才能保證不管誰在哪裡打開都是同一種格式。"""
    now = datetime.now(TAIPEI_TZ) if TAIPEI_TZ else datetime.now()
    return f"{now.year}/{now.month}/{now.day} {now.hour:02d}:{now.minute:02d}"


def safe_numeric(v):
    """安全轉換為數字，供報表格式化前使用。萬一因為未預期的資料狀況 (例如
    來源檔案有特殊欄位命名或型別問題) 讓某個儲存格的值不是單純的數字，
    也不應該讓整份報表直接崩潰 —— 寧可退回顯示原樣文字，事後方便追查，
    也不要讓使用者完全看不到報表。回傳 (數值或None, 是否為合法數字)。"""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if pd.isna(v):
                return 0.0, True
        except (TypeError, ValueError):
            pass
        return float(v), True
    try:
        n = pd.to_numeric(v)
        return float(n), True
    except (TypeError, ValueError):
        return v, False


# ============================================================
# 樣式化 HTML 預覽 (比照原系統的綠底標題、小計/總計配色)
# ============================================================

HTML_FONT_FAMILY = "'Microsoft JhengHei', 'Noto Sans TC', 'PingFang TC', 'Heiti TC', 'Microsoft YaHei', Arial, sans-serif"


def build_html_table(rows, row_fields, value_cols, pct_cols, growth_cols, report_title, summary_html=""):
    extra_cols = set(pct_cols + growth_cols)
    display_cols = order_display_cols(value_cols, pct_cols, growth_cols)
    headers = row_fields + display_cols

    font_family = HTML_FONT_FAMILY
    th_style = f"color:#FFFFFF;background-color:#00695C;text-align:center;border:1px solid #D9D9D9;padding:8px 10px;font-weight:bold;font-family:{font_family};"
    td_style = f"border:1px solid #D9D9D9;padding:6px 10px;font-family:{font_family};"

    html = [
        f"<div style='background:#FFFFFF;border-radius:10px;padding:18px;border:1px solid #eee;font-family:{font_family};'>",
        f"<h3 style='text-align:center;color:#004D40;margin-top:0;font-family:{font_family};'>{report_title}</h3>",
        summary_html,
        "<div style='overflow-x:auto;'>",
        f"<table id='report-table' style='border-collapse:separate;border-spacing:0;width:100%;font-size:14px;font-family:{font_family};'>",
        "<tr>",
    ]
    for f in row_fields:
        html.append(f"<th style='{th_style}'>{f}</th>")
    for c in display_cols:
        year, label = pretty_header(c)
        if year:
            html.append(f"<th style='{th_style}'>{year}<br>{label}</th>")
        else:
            html.append(f"<th style='{th_style}'>{label}</th>")
    html.append("</tr>")

    for r in rows:
        bg = ""
        fw = "font-weight:normal;"
        if r["type"] == "subtotal":
            bg = "background-color:#E0F2F1;"
            fw = "font-weight:bold;"
        elif r["type"] == "total":
            bg = "background-color:#B2DFDB;"
            fw = "font-weight:bold;"
        html.append(f"<tr style='{bg}{fw}'>")
        for i, f in enumerate(row_fields):
            v = r["values"].get(f, "")
            align = "left" if i == 0 else "center"
            html.append(f"<td style='{td_style}text-align:{align};'>{v}</td>")
        overrides = r.get("overrides", {})
        for c in display_cols:
            if c in extra_cols:
                v, ok = safe_numeric(r["sums"].get(c, 0))
                html.append(f"<td style='{td_style}text-align:right;'>{v:.1%}</td>" if ok else f"<td style='{td_style}text-align:right;color:#999999;'>{v}</td>")
            elif c in overrides:
                html.append(f"<td style='{td_style}text-align:right;color:#999999;'>{overrides[c]}</td>")
            else:
                raw = r["sums"].get(c, "")
                if raw == "":
                    html.append(f"<td style='{td_style}'></td>")
                else:
                    v, ok = safe_numeric(raw)
                    html.append(f"<td style='{td_style}text-align:right;'>{v:,.0f}</td>" if ok else f"<td style='{td_style}text-align:right;color:#999999;'>{v}</td>")
        html.append("</tr>")

    html.append("</table></div>")
    footer_style = f"display:flex;justify-content:space-between;margin-top:8px;font-size:12px;color:#333333;font-family:{font_family};"
    html.append(
        f"<div style='{footer_style}'><span>{get_report_timestamp()}</span><span>第 1 頁</span></div>"
    )
    html.append("</div>")
    return "".join(html)


CAPTURE_HTML_TEMPLATE = """
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js"></script>
<div style="text-align:center;">
  <button id="export-png-btn-{uid}" style="background-color:#00695C;color:white;border:none;padding:10px 20px;
    border-radius:6px;font-size:14px;cursor:pointer;width:100%;">🖼️ 匯出 PNG 圖片</button>
</div>
<div id="result-area-{uid}" style="margin-top:10px;font-size:13px;color:#00695C;text-align:center;"></div>
<div id="capture-wrap-{uid}" style="position:absolute; left:-99999px; top:0; width:max-content;">{table_html}</div>
<script>
(function() {{
    const resultArea = document.getElementById('result-area-{uid}');
    const pngBtn = document.getElementById('export-png-btn-{uid}');
    let busy = false;

    async function captureCanvas() {{
        const target = document.getElementById('capture-wrap-{uid}');
        // 等待字型完全載入後才擷取，避免在字型還沒套用完成時就截圖，導致回退成瀏覽器預設的 serif 字體
        if (document.fonts && document.fonts.ready) {{
            await document.fonts.ready;
        }}
        if (typeof html2canvas === 'undefined') {{
            throw new Error('html2canvas 尚未載入完成，請確認網路連線後再試一次');
        }}
        return await html2canvas(target, {{
            scale: 4,
            backgroundColor: '#ffffff',
            windowWidth: target.scrollWidth,
            windowHeight: target.scrollHeight,
            width: target.scrollWidth,
            height: target.scrollHeight,
            useCORS: true
        }});
    }}

    async function shareOrDownload(blob, filename, mime) {{
        const file = new File([blob], filename, {{type: mime}});
        const isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent) ||
            (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);

        if (isMobile) {{
            if (navigator.canShare && navigator.canShare({{files: [file]}})) {{
                // 注意：navigator.share 若同時帶入 files 與 title/text，iOS 選「儲存到檔案」
                // 時會把 title 文字也另存成一個獨立的小型文字檔，因此這裡只傳 files。
                try {{
                    await navigator.share({{files: [file]}});
                    resultArea.innerText = '✅ 已開啟分享選單';
                }} catch (shareErr) {{
                    // AbortError = 使用者按了取消，不視為錯誤
                }}
            }} else {{
                const url = URL.createObjectURL(blob);
                resultArea.innerHTML =
                    "目前瀏覽器不支援原生分享，請長按下方圖片選擇「儲存影像」：<br/>" +
                    "<img src='" + url + "' style='max-width:100%;border-radius:8px;border:1px solid #eee;margin-top:6px;' />";
            }}
        }} else {{
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = filename;
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            setTimeout(function() {{ URL.revokeObjectURL(url); }}, 5000);
            resultArea.innerText = '✅ 已開始下載';
        }}
    }}

    pngBtn.addEventListener('click', async function() {{
        if (busy) return;
        busy = true; pngBtn.disabled = true;
        resultArea.innerText = '🔄 產生圖片中，請稍候...';
        try {{
            const canvas = await captureCanvas();
            const blob = await new Promise(function(resolve) {{ canvas.toBlob(resolve, 'image/png'); }});
            if (!blob) {{ throw new Error('圖片產生失敗 (canvas 轉換為空)'); }}
            await shareOrDownload(blob, '{filename}.png', 'image/png');
        }} catch (err) {{
            resultArea.innerText = '❌ 匯出失敗：' + (err && err.message ? err.message : err);
        }} finally {{
            busy = false; pngBtn.disabled = false;
        }}
    }});
}})();
</script>
"""


EXCEL_SHARE_HTML_TEMPLATE = """
<div style="text-align:center;">
  <button id="excel-share-btn-{uid}" style="background-color:#00695C;color:white;border:none;padding:10px 20px;
    border-radius:6px;font-size:14px;cursor:pointer;width:100%;">📄 下載 / 分享 Excel 報表</button>
</div>
<script>
(function() {{
    const btn = document.getElementById('excel-share-btn-{uid}');
    let busy = false;  // 防止使用者連點，或分享流程尚未結束前重複觸發，導致產生兩個檔案
    btn.addEventListener('click', async function() {{
        if (busy) return;
        busy = true;
        btn.disabled = true;
        try {{
            const b64 = "{b64data}";
            const byteChars = atob(b64);
            const byteNumbers = new Array(byteChars.length);
            for (let i = 0; i < byteChars.length; i++) {{
                byteNumbers[i] = byteChars.charCodeAt(i);
            }}
            const byteArray = new Uint8Array(byteNumbers);
            const mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet';
            const blob = new Blob([byteArray], {{type: mime}});
            const file = new File([blob], '{filename}.xlsx', {{type: mime}});
            const isMobile = /iPhone|iPad|iPod|Android/i.test(navigator.userAgent) ||
                (navigator.platform === 'MacIntel' && navigator.maxTouchPoints > 1);

            if (isMobile && navigator.canShare && navigator.canShare({{files: [file]}})) {{
                // 用系統分享面板取代 <a download>，避免 iOS 加入主畫面的 App
                // 因為缺少瀏覽器介面（無上一頁鍵）而卡在 Quick Look 檔案預覽畫面出不去。
                // 注意：若同時帶入 files 與 title，iOS 選「儲存到檔案」時會把 title 文字
                // 也另存成一個獨立的小型文字檔 (曾實際出現 47 byte 的「文字」檔)，
                // 因此這裡只傳 files，不帶 title。
                // 使用者若在分享面板中直接取消，不做任何提示、也不會留下任何檔案。
                try {{
                    await navigator.share({{files: [file]}});
                }} catch (shareErr) {{
                    // AbortError = 使用者按了取消，不需視為錯誤、也不顯示任何訊息
                }}
            }} else {{
                const url = URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = '{filename}.xlsx';
                document.body.appendChild(a);
                a.click();
                document.body.removeChild(a);
                setTimeout(function() {{ URL.revokeObjectURL(url); }}, 5000);
            }}
        }} finally {{
            busy = false;
            btn.disabled = false;
        }}
    }});
}})();
</script>
"""


def render_excel_share(excel_bytes: bytes, filename: str, uid: str):
    """以分享面板 (navigator.share) 方式提供 Excel 下載，避免 iOS 加入主畫面
    (standalone PWA) 因缺少瀏覽器導覽列，下載後卡在 Quick Look 檔案預覽頁面
    卻無法返回 App 的問題。不顯示任何下載/分享狀態文字；並以 busy flag 防止
    重複觸發造成產生兩個檔案。"""
    b64data = base64.b64encode(excel_bytes).decode()
    safe_uid = re.sub(r"[^0-9A-Za-z_]", "_", uid)
    components.html(
        EXCEL_SHARE_HTML_TEMPLATE.format(b64data=b64data, filename=filename, uid=safe_uid),
        height=50,
    )


# ============================================================
# Excel 匯出 (比照原系統的綠底標題、A4 橫式、標題列重複列印)
# ============================================================

def generate_excel_bytes(rows, row_fields, value_cols, pct_cols, growth_cols, report_title, summary_lines=None):
    extra_cols = set(pct_cols + growth_cols)
    display_cols = order_display_cols(value_cols, pct_cols, growth_cols)
    headers = row_fields + display_cols

    wb = Workbook()
    ws = wb.active
    ws.title = "分析報表"

    ws.views.sheetView[0].showGridLines = False
    ws.views.sheetView[0].zoomScale = 85
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True

    # 列印邊界上下左右統一 1.2 公分 (openpyxl 的 PageMargins 單位是英吋，1.2cm ÷ 2.54 ≈ 0.4724 吋)
    margin_cm_in_inch = 1.2 / 2.54
    ws.page_margins = PageMargins(
        left=margin_cm_in_inch, right=margin_cm_in_inch,
        top=margin_cm_in_inch, bottom=margin_cm_in_inch,
        header=0, footer=0,
    )

    # 頁尾：左下角顯示日期＋時間 (報表產出當下的固定文字，24小時制、不含AM/PM，
    # 不用 Excel 的 &D/&T 動態欄位代碼，因為那會依開啟者當下的 Excel 地區設定顯示成
    # 不同格式，例如跑出12小時制加AM/PM)；右下角顯示頁碼 (&P 是 Excel 內建的頁碼代碼，
    # 這個會自動對應實際列印頁數，維持動態即可)。
    # &"微軟正黑體,標準"&11 是 Excel 頁首頁尾專用的字型控制代碼，讓頁尾文字跟表格內容
    # 使用同一種字型 (微軟正黑體)、同樣大小 (11号)，視覺上一致。注意：字型大小代碼 &11
    # 後面刻意保留一個空格，避免緊接著的日期文字開頭數字 (例如"2026") 被 Excel 誤判
    # 合併成 "&112026" (當成字型大小112026號去解析)，導致文字顯示被吃掉一部分。
    footer_font_code = '&"微軟正黑體,標準"&11 '
    ws.oddFooter.left.text = f"{footer_font_code}{get_report_timestamp()}"
    ws.oddFooter.right.text = f"{footer_font_code}第 &P 頁"

    header_fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
    subtotal_fill = PatternFill(start_color="E0F2F1", end_color="E0F2F1", fill_type="solid")
    total_fill = PatternFill(start_color="B2DFDB", end_color="B2DFDB", fill_type="solid")
    font_title = Font(name="微軟正黑體", bold=True, size=16, color="004D40")
    font_b_w = Font(name="微軟正黑體", bold=True, color="FFFFFF", size=12)
    font_norm = Font(name="微軟正黑體", size=11)
    font_bold = Font(name="微軟正黑體", bold=True, size=11)
    font_summary = Font(name="微軟正黑體", size=13, bold=True, color="333333")
    font_rate = Font(name="微軟正黑體", size=15, bold=True, color="C00000")
    align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_r = Alignment(horizontal="right", vertical="center")
    border_thin = Border(*[Side(style="thin", color="D9D9D9")] * 4)
    header_border = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    ws.cell(row=1, column=1, value=report_title).font = font_title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    next_row = 3
    if summary_lines:
        for label, value, fmt in summary_lines:
            ws.cell(row=next_row, column=1, value=label)
            if value is not None:
                cell = ws.cell(row=next_row, column=2, value=value)
                if fmt:
                    cell.number_format = fmt
                    cell.font = font_rate if fmt == "0.00%" else font_summary
            ws.cell(row=next_row, column=1).font = font_rate if (fmt == "0.00%") else font_summary
            next_row += 1
        next_row += 1  # 空一行

    header_row = next_row
    ws.append([])

    header_texts = []
    for f in row_fields:
        header_texts.append(f)
    for c in display_cols:
        year, label = pretty_header(c)
        header_texts.append(f"{year}\n{label}" if year else label)
    ws.append(header_texts)
    ws.row_dimensions[header_row].height = 48
    ws.print_title_rows = f"1:{header_row}"

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = font_b_w
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = align_c

    current_row = header_row + 1
    for r in rows:
        overrides = r.get("overrides", {})
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i)
            if h in value_cols and h in overrides:
                # 明細列若原始資料是「無資料(-)」或異常標記文字 (例如 #N/A、Unknown)，
                # 保留原樣文字顯示，跟「真正的0」做出區別；小計/總計不受影響，仍是加總後的數字
                cell.value = overrides[h]
                cell.alignment = align_r
                cell.font = Font(name="微軟正黑體", size=11, color="999999")
            elif h in value_cols:
                raw = r["sums"].get(h, "")
                v, ok = safe_numeric(raw)
                if ok:
                    cell.value = v
                    cell.number_format = "#,##0"
                else:
                    cell.value = str(v)
                    cell.font = Font(name="微軟正黑體", size=11, color="999999")
                cell.alignment = align_r
                if ok:
                    cell.font = font_bold if r["type"] != "data" else font_norm
            elif h in extra_cols:
                raw = r["sums"].get(h, 0)
                v, ok = safe_numeric(raw)
                if ok:
                    cell.value = v
                    cell.number_format = "0.0%"
                else:
                    cell.value = str(v)
                    cell.font = Font(name="微軟正黑體", size=11, color="999999")
                cell.alignment = align_r
                if ok:
                    cell.font = font_bold if r["type"] != "data" else font_norm
            else:
                cell.value = r["values"].get(h, "")
                cell.alignment = Alignment(horizontal="left" if i == 1 else "center", vertical="center", wrap_text=True)
                cell.font = font_bold if r["type"] != "data" else font_norm
            cell.border = border_thin
            if r["type"] == "subtotal":
                cell.fill = subtotal_fill
            elif r["type"] == "total":
                cell.fill = total_fill
        ws.row_dimensions[current_row].height = 42
        current_row += 1

    for col_idx, col_name in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = 28
        else:
            ws.column_dimensions[col_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ============================================================
# 處方釋出率分析（固定格式）
# ============================================================

def analysis_prescription(df_filtered, vendor_name=None, show_vendor=False):
    if df_filtered.empty:
        return None
    ds_hp_cond = (df_filtered["通路"] == "DS") & (df_filtered["層級別"].isin(["1.醫學中心", "2.區域醫院", "3.地區醫院"]))
    ds_hp_val = df_filtered[ds_hp_cond]["2024年申報量(顆)"].sum()
    hp_val = df_filtered[df_filtered["通路"] == "HP"]["2024年申報量(顆)"].sum()
    total_val = ds_hp_val + hp_val
    rate = (ds_hp_val / total_val * 100) if total_val > 0 else 0
    title = f"{vendor_name}_醫院處方釋出率(2024年)" if vendor_name else "整體醫院處方釋出率(2024年)"

    row_fields = ["成分簡稱", "單複方", "劑型", "含量"]
    if show_vendor and "廠商簡稱" in df_filtered.columns:
        row_fields.append("廠商簡稱")
    row_fields += ["通路", "層級別"]
    row_fields = [f for f in row_fields if f in df_filtered.columns]
    value_cols = [c for c in ["2022年申報量(顆)", "2023年申報量(顆)", "2024年申報量(顆)"] if c in df_filtered.columns]

    rows, pct_cols, growth_cols = build_nested_rows(df_filtered, row_fields, ["通路"], value_cols, pct_years=[], add_growth=False)

    return {
        "title": title, "ds_hp_val": ds_hp_val, "hp_val": hp_val, "total_val": total_val, "rate": rate,
        "rows": rows, "row_fields": row_fields, "value_cols": value_cols, "pct_cols": pct_cols, "growth_cols": growth_cols,
    }


# ============================================================
# Streamlit 主畫面
# ============================================================

st.title("📊 華典資料庫分析系統")

ANALYSIS_TO_SOURCE = {
    "1. 廠商分析": "data_hp_gp_ds.csv",
    "2. 層級別分析": "data_hp_gp_ds.csv",
    "3. 科別分析": "data_department.csv",
    "4. 推估醫院分析": "data_hospital.csv",
}
ANA_OPTIONS = list(ANALYSIS_TO_SOURCE.keys()) + ["5. 處方釋出率分析"]

ana_choice = st.selectbox("第1步：選擇分析功能", ANA_OPTIONS, key="ana_choice")

# ------------------------------------------------------------
# 1~4：自訂樞紐分析
# ------------------------------------------------------------
if ana_choice in ANALYSIS_TO_SOURCE:
    st.caption("先選擇成分，接著把需要的欄位拖到「報表欄位」，即時預覽會隨著您的設定更新，效果如同 Excel 樞紐分析。")

    source_path = ANALYSIS_TO_SOURCE[ana_choice]
    df_raw = load_data(source_path)

    if df_raw.empty:
        st.warning(f"找不到資料檔案「{source_path}」，請確認檔案已放置於工作目錄。")
    elif "成分簡稱" not in df_raw.columns:
        st.error(f"資料檔案缺少「成分簡稱」欄位，實際欄位為：{list(df_raw.columns)}")
    else:
        comp_options = sorted([c for c in df_raw["成分簡稱"].dropna().unique() if c])
        # 單一元件即支援輸入時自動篩選（Streamlit multiselect 內建 type-to-search）
        comps_selected = st.multiselect(
            "第2步：輸入關鍵字並選擇成分品項 (必選)",
            options=comp_options,
            key=f"comps_{ana_choice}",
            placeholder="輸入關鍵字，例如 Levofloxacin",
        )

        if not comps_selected:
            st.info("請先選擇至少一個成分，才會顯示後續的欄位設定。")
        else:
            df_comp = df_raw[df_raw["成分簡稱"].isin(comps_selected)]

            # __mk 結尾的欄位是內部用來記錄「無資料(-)/異常標記(#N/A、Unknown)」的輔助欄位
            # (詳見 load_data)，純粹供報表渲染時查詢使用，不應該出現在使用者可選的欄位清單裡
            other_cols = [c for c in df_raw.columns if c != "成分簡稱" and not c.endswith("__mk")]
            value_cols_auto = [c for c in other_cols if ("申報量" in c) or ("金額" in c)]
            dim_cols_all = [c for c in other_cols if c not in value_cols_auto]

            st.markdown("### 🧩 第3步：拖曳欄位到「報表欄位」，並排序")
            dnd_key = f"dnd_{ana_choice}_{'_'.join(sorted(comps_selected))}"

            state_key = f"pivot_state_{dnd_key}"
            if state_key not in st.session_state:
                st.session_state[state_key] = {"available": dim_cols_all, "selected": []}
            else:
                # 換了成分/分析功能等情境下，欄位清單內容可能改變，這裡做防呆同步
                prev = st.session_state[state_key]
                avail = [c for c in prev["available"] if c in dim_cols_all]
                sel = [c for c in prev["selected"] if c in dim_cols_all]
                known = set(avail) | set(sel)
                avail += [c for c in dim_cols_all if c not in known]
                st.session_state[state_key] = {"available": avail, "selected": sel}

            # 拖曳排序元件 (streamlit-sortables) 是第三方套件，偶爾會在拖曳當下丟出
            # 「Minified React error #185」(該元件內部的無限更新循環，屬於它自己的
            # 相容性問題，並非資料或計算邏輯出錯)。提供一個手動切換開關，萬一拖曳
            # 元件又出現這個錯誤，可以立即改用不會壞掉的下拉式勾選繼續操作，
            # 不必等待這個第三方套件修好。勾選的順序即為欄位排列順序。
            use_fallback = st.checkbox(
                "⚠️ 拖曳排序出現錯誤時，改用勾選方式選擇欄位 (勾選順序＝欄位順序)",
                value=False, key=f"use_fallback_{dnd_key}",
            )

            if HAS_SORTABLES and not use_fallback:
                containers = sort_items(
                    [
                        {"header": "📋 可用欄位 (拖曳到下方使用)", "items": st.session_state[state_key]["available"]},
                        {"header": "📊 報表欄位 (由左到右排列；如同樞紐分析表的欄位區)", "items": st.session_state[state_key]["selected"]},
                    ],
                    multi_containers=True,
                    direction="horizontal",
                    key=dnd_key,
                )
                if containers and len(containers) > 1:
                    st.session_state[state_key] = {"available": containers[0]["items"], "selected": containers[1]["items"]}
                row_fields = st.session_state[state_key]["selected"]
            elif HAS_SORTABLES:
                # 使用者主動切換成勾選模式：以 multiselect 取代拖曳，選取順序即欄位順序，
                # 並同步回 session_state，切回拖曳模式時能延續同樣的欄位與順序
                sel = st.multiselect(
                    "選擇報表欄位 (依勾選順序排列)",
                    options=dim_cols_all,
                    default=st.session_state[state_key]["selected"],
                    key=f"fallback_fields_{dnd_key}",
                )
                st.session_state[state_key] = {
                    "available": [c for c in dim_cols_all if c not in sel],
                    "selected": sel,
                }
                row_fields = sel
            else:
                st.error("尚未安裝 streamlit-sortables 套件，暫以勾選方式呈現，請於 requirements.txt 加入 streamlit-sortables 後重新部署即可拖曳。")
                row_fields = st.multiselect("選擇報表欄位", options=dim_cols_all, key=f"fallback_fields_{ana_choice}")

            if not row_fields:
                st.info("請至少拖曳一個欄位到「報表欄位」區塊。")
            else:
                st.markdown("### 🔍 第4步：逐欄篩選與合計 (點欄位下方的「🔽 篩選 / 合計」展開設定；如同 Excel 欄篩選)")
                st.caption("💡 篩選會彼此連動：某欄位選擇後，其他欄位只會列出仍有對應資料的選項，避免篩出不存在的組合。")

                # 先讀取目前各欄位已選的篩選值 (用於彼此連動縮小可選範圍)
                current_filters = {}
                for f in row_fields:
                    key = f"filt_{dnd_key}_{f}"
                    if key in st.session_state and st.session_state[key]:
                        current_filters[f] = st.session_state[key]

                filter_ui_cols = st.columns(len(row_fields))
                filters = {}
                subtotal_fields_selected = []
                for i, f in enumerate(row_fields):
                    with filter_ui_cols[i]:
                        st.markdown(f"**{f}**")
                        with st.expander("🔽 篩選 / 合計", expanded=False):
                            # 依其他欄位目前的篩選結果，動態縮小此欄位可選項目
                            df_scope = df_comp
                            for other_col, other_sel in current_filters.items():
                                if other_col != f and other_sel:
                                    df_scope = df_scope[df_scope[other_col].isin(other_sel)]
                            options = sorted([v for v in df_scope[f].dropna().unique() if str(v).strip() != ""])

                            key = f"filt_{dnd_key}_{f}"
                            # 若其他欄位變動導致此欄位先前選的值已不存在，先清掉避免元件報錯
                            if key in st.session_state:
                                st.session_state[key] = [v for v in st.session_state[key] if v in options]

                            sel = st.multiselect(f"篩選「{f}」(不選代表全選)", options=options, key=key)
                            if sel:
                                filters[f] = sel
                            if st.checkbox(f"Σ 此欄要合計", key=f"sub_{dnd_key}_{f}"):
                                subtotal_fields_selected.append(f)
                subtotal_fields = [f for f in row_fields if f in subtotal_fields_selected]

                df_filtered = df_comp.copy()
                for col, sel in filters.items():
                    df_filtered = df_filtered[df_filtered[col].isin(sel)]

                st.caption(f"篩選後共 **{len(df_filtered):,}** 筆資料")

                c_val1, c_val2 = st.columns([2, 1])
                with c_val1:
                    value_cols = st.multiselect(
                        "第5步：選擇要加總的數值欄位 (預設帶入所有申報量／金額欄位)",
                        options=value_cols_auto, default=value_cols_auto, key=f"vals_{dnd_key}",
                    )
                with c_val2:
                    qty_years_avail = sorted(set(c[:4] for c in value_cols if "申報量" in c))
                    has_qty = len(qty_years_avail) > 0
                    pct_years = st.multiselect(
                        "➕ 加入年度占比(%) (可只選需要的年份)",
                        options=qty_years_avail, default=[], disabled=not has_qty, key=f"pct_{dnd_key}",
                    )
                    add_growth = st.checkbox("➕ 加入年度成長率(%)", value=False, disabled=not has_qty, key=f"growth_{dnd_key}")

                # Req 7：檔名納入成分與所有篩選項目
                filename_parts = list(comps_selected)
                for col in row_fields:
                    if col in filters:
                        filename_parts.append("_".join(filters[col]))
                filename_parts.append(ana_choice.split(".")[1].strip())
                report_title = "_".join(filename_parts)

                full_row_fields = ["成分簡稱"] + row_fields

                if value_cols and not df_filtered.empty:
                    rows, pct_cols, growth_cols = build_nested_rows(
                        df_filtered, full_row_fields, subtotal_fields, value_cols, pct_years, add_growth
                    )
                    st.markdown("### 📄 報表即時預覽")
                    table_html = build_html_table(rows, full_row_fields, value_cols, pct_cols, growth_cols, report_title)
                    st.markdown(table_html, unsafe_allow_html=True)

                    st.markdown("### 📥 下載")
                    dl_col1, dl_col2 = st.columns(2)
                    safe_dnd_key = re.sub(r"[^0-9A-Za-z_]", "_", dnd_key)
                    with dl_col1:
                        excel_bytes = generate_excel_bytes(rows, full_row_fields, value_cols, pct_cols, growth_cols, report_title)
                        render_excel_share(excel_bytes, report_title, uid=f"pivot_{dnd_key}")
                    with dl_col2:
                        # Req 6：不再顯示第二份預覽，表格離屏渲染僅供 html2canvas 擷取；
                        # Req 5：windowWidth/Height 依內容實際尺寸擷取，避免手機/小視窗被裁切
                        components.html(
                            CAPTURE_HTML_TEMPLATE.format(table_html=table_html, filename=report_title, uid=f"pivot_{safe_dnd_key}"),
                            height=90,
                        )
                elif not value_cols:
                    st.warning("⚠️ 請至少選擇一個要加總的數值欄位。")
                else:
                    st.warning("❌ 篩選後無資料，請放寬篩選條件。")

# ------------------------------------------------------------
# 5：處方釋出率分析（固定格式）
# ------------------------------------------------------------
else:
    st.caption("分析成分從醫院端流向藥局處方的釋出率，可看單一廠商。此分析維持固定格式，不提供自訂欄位。")

    df_hp = load_data("data_hp_gp_ds.csv")
    if df_hp.empty:
        st.warning("找不到資料檔案「data_hp_gp_ds.csv」，請確認檔案已放置於工作目錄。")
    elif "成分簡稱" not in df_hp.columns:
        st.error(f"資料檔案缺少「成分簡稱」欄位，實際欄位為：{list(df_hp.columns)}")
    else:
        comp_options = sorted([c for c in df_hp["成分簡稱"].dropna().unique() if c])
        comp = st.selectbox(
            "第2步：輸入關鍵字並選擇單一成分品項 (必選)",
            options=[""] + comp_options, key="presc_comp",
        )

        combo, form, dose, vendor = None, None, None, None
        if comp:
            df_c = df_hp[df_hp["成分簡稱"] == comp]
            combo_options = sorted([v for v in df_c["單複方"].dropna().unique() if v]) if "單複方" in df_c.columns else []
            combo = st.selectbox("第3步：選擇單一單複方 (必選)", options=[""] + combo_options, key="presc_combo")

            if combo:
                df_cb = df_c[df_c["單複方"] == combo]
                form_options = sorted([v for v in df_cb["劑型"].dropna().unique() if v])
                form = st.selectbox("第4步：選擇單一劑型 (必選)", options=[""] + form_options, key="presc_form")

                if form:
                    df_cbf = df_cb[df_cb["劑型"] == form]
                    dose_options = sorted([v for v in df_cbf["含量"].dropna().unique() if v], key=parse_dosage)
                    dose = st.selectbox("第5步：選擇單一含量 (必選)", options=[""] + dose_options, key="presc_dose")

                    if dose:
                        df_cbfd = df_cbf[df_cbf["含量"] == dose]
                        show_vendor = st.checkbox("若需顯示單一廠商，請勾選", key="presc_show_vendor")
                        if show_vendor:
                            vendor_options = sorted([v for v in df_cbfd["廠商簡稱"].dropna().unique() if v])
                            vendor = st.selectbox("第6步：選擇單一廠商 (必選)", options=[""] + vendor_options, key="presc_vendor")

        st.divider()
        if st.button("🚀 產生報表", type="primary", key="presc_generate"):
            missing = []
            if not comp: missing.append("成分")
            if not combo: missing.append("單複方")
            if not form: missing.append("劑型")
            if not dose: missing.append("含量")
            if st.session_state.get("presc_show_vendor") and not vendor:
                missing.append("廠商")

            if missing:
                st.error(f"⚠️ 請務必選擇：{'、'.join(missing)}！")
            else:
                df_f = df_hp[
                    (df_hp["成分簡稱"] == comp) & (df_hp["單複方"] == combo) &
                    (df_hp["劑型"] == form) & (df_hp["含量"] == dose)
                ]
                if vendor:
                    df_f = df_f[df_f["廠商簡稱"] == vendor]

                summary = analysis_prescription(df_f, vendor, bool(st.session_state.get("presc_show_vendor")))
                if not summary:
                    st.error("❌ 找不到符合條件的資料。")
                else:
                    filename_parts = [comp, combo, form, dose]
                    if vendor:
                        filename_parts.append(vendor)
                    filename_parts.append("醫院處方釋出率分析")
                    report_filename = "_".join(filename_parts)

                    summary_box_html = f"""
                        <div style='background-color:#E0F2F1; padding:20px; border-radius:10px;
                            border-left: 6px solid #00695C; margin-bottom:16px;'>
                            <ul style='font-size:18px; color:#004D40; list-style-type:none; padding-left:0; line-height:1.6; margin:0;'>
                                <li><b>DS from HP：</b> {summary['ds_hp_val']:,.0f}</li>
                                <li><b>HP：</b> {summary['hp_val']:,.0f}</li>
                                <li><b>DS from HP + HP：</b> {summary['total_val']:,.0f}</li>
                                <li style='margin-top:6px; font-size:13px; color:#00695C;'>
                                    公式：處方釋出率 ＝ DS from HP ÷ (DS from HP + HP) × 100%</li>
                                <li style='margin-top:10px; font-size:26px;'><b>{summary['title']}：</b>
                                    <span style='color:#D32F2F; font-weight:bold;'>{summary['rate']:.2f}%</span></li>
                            </ul>
                        </div>
                        """

                    st.markdown("### 📄 報表即時預覽")
                    table_html = build_html_table(
                        summary["rows"], summary["row_fields"], summary["value_cols"],
                        summary["pct_cols"], summary["growth_cols"], report_filename, summary_box_html,
                    )
                    st.markdown(table_html, unsafe_allow_html=True)

                    st.markdown("### 📥 下載")
                    dl_col1, dl_col2 = st.columns(2)
                    with dl_col1:
                        summary_lines = [
                            ("DS from HP:", summary["ds_hp_val"], "#,##0"),
                            ("HP:", summary["hp_val"], "#,##0"),
                            ("DS from HP + HP:", summary["total_val"], "#,##0"),
                            ("公式：處方釋出率 ＝ DS from HP ÷ (DS from HP + HP) × 100%", None, None),
                            (f"{summary['title']}:", summary["rate"] / 100, "0.00%"),
                        ]
                        excel_bytes = generate_excel_bytes(
                            summary["rows"], summary["row_fields"], summary["value_cols"],
                            summary["pct_cols"], summary["growth_cols"], report_filename, summary_lines,
                        )
                        render_excel_share(excel_bytes, report_filename, uid="presc")
                    with dl_col2:
                        components.html(
                            CAPTURE_HTML_TEMPLATE.format(table_html=table_html, filename=report_filename, uid="presc"),
                            height=90,
                        )
